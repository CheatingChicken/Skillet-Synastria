import openpyxl

wb = openpyxl.load_workbook(r'g:\Programmieren\WoW Addons\Synastria\Skillet - Synastria\Planning\Skillet\Scoots ID Upgradables.xlsx')

# Extract all upgrades with item names from all sheets
all_upgrades = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    for row in ws.iter_rows(values_only=True):
        if len(row) >= 5 and row[0] is not None and row[1] is not None and row[3] is not None and row[4] is not None:
            try:
                source_name = str(row[0]).strip()
                source_id = int(float(row[1]))
                target_name = str(row[3]).strip()
                target_id = int(float(row[4]))
                if source_id and target_id and source_name and target_name:
                    all_upgrades[source_id] = (target_id, source_name, target_name)
            except:
                pass

# Print Lua format with comments
for source in sorted(all_upgrades.keys()):
    target_id, source_name, target_name = all_upgrades[source]
    print(f"    [{source}] = {target_id},  -- {source_name} -> {target_name}")
